
import os
from django.http import JsonResponse
from django.shortcuts import redirect, render,HttpResponse
from django.views.decorators.csrf import csrf_exempt

from user_manage import models



def Classified_storage(file_object):
    import os
 
    if file_object.content_type.startswith('image'):
        root_dir = 'user_manage/static/open_upload/img/'
    elif file_object.content_type.startswith('video'):
        root_dir = 'user_manage/static/open_upload/video/'
    elif file_object.content_type.startswith('audio'):
        root_dir = 'user_manage/static/open_upload/audio/'
    else:
        root_dir = 'user_manage/static/open_upload/others/'
    
    if not os.path.exists(root_dir):
        os.makedirs(root_dir)
    
    filepath = os.path.join(root_dir, file_object.name)
    
    return filepath
# 第一部分 通过 open 获取存储文件
def upload_list(request):

    if request.method=='GET':
        return render(request, 'upload_list.html')
    
    print(request.POST) #请求中的数据
    print(request.FILES) # 请求发送过来的文件
    file_object = request.FILES.get('avatar')
    print(file_object.name) # 打印文件名


    file_obj=Classified_storage(file_object) # 判断并创建文件目录

    with open (file_obj ,mode='wb') as f:
        for chunk in file_object.chunks():
            f.write(chunk)
            # print(chunk)
        

    return HttpResponse('添加成功')




def upload_multi(request):
    """ 批量删除（Excel文件）"""
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')




from django import forms
from user_manage.utils.bootstrap import BootStrapForm, BootStrapModelForm


class UpForm(BootStrapForm):
    bootstrap_exclude_fields = ['img']

    depart = forms.CharField(label="部门")
    sales = forms.IntegerField(label="销售额")
    file = forms.FileField(label="文件")

def upload_form(request):

    '''基于form做文件上级'''
    title = "Form上传"
    if request.method == "GET":
        form = UpForm()
        return render(request, 'upload_form.html', {"form": form, "title": title})

    form = UpForm(data=request.POST, files=request.FILES)
    if form.is_valid():
       
        # 1.读取图片内容，写入到文件夹中并获取文件的路径。
        image_object = form.cleaned_data.get("file")

        db_fileform_path = os.path.join('static','img' ,image_object.name) #如果需要可以将存储到数据的的单独拆分 
        fileform_path = os.path.join("user_manage", db_fileform_path)

        # 基于media 作为添加路径
        # fileform_path = os.path.join("media", image_object.name)
        f = open(fileform_path, mode='wb')
        for chunk in image_object.chunks():
            f.write(chunk)
        f.close()

        # 2.将内容文件路径写入到数据库
        models.performance.objects.create(
            depart=form.cleaned_data['depart'],
            sales=form.cleaned_data['sales'],
            file=fileform_path,
        )
        return redirect('/upload/form/')
    return render(request, 'upload_form.html', {"form": form, "title": title})


from django import forms
from user_manage.utils.form import BootStrapModelForm


class UpModelForm(BootStrapModelForm):

    # img 不需要配置默认样式
    bootstrap_exclude_fields = ['img']

    class Meta:
        model =models.Picture
        fields = "__all__"

def upload_model_form_list(request):
    queryset = models.Picture.objects.all()
    return render(request, 'upload_model_form_list.html', {'queryset': queryset})

def upload_model_form_add(request):
    '''上传文件和数据(modelForm)'''
    title ='ModelForm上传文件'
    if request.method=='GET':
        form = UpModelForm()
        

        return render(request,'upload_model_form_add.html',{'form':form, 'title':title })
        
    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        # 对于文件：自动保存；
        # 字段 + 上传路径写入到数据库
        form.save()
        return redirect("/upload/model_form/list")
    return render(request, 'upload_model_form_add.html', {"form": form, 'title': title})

    